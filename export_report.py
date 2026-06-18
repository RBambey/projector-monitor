"""Generate an Excel report with line chart from today's temperature log CSV."""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Series colours matching the app's DSKY palette
_SERIES_COLORS = ["39FF14", "00D4FF", "FF8C00", "CC88FF", "FFDD00",
                  "FF2222", "FFFFFF", "00FF88"]


def generate_report(log_dir: Path, threshold: float) -> Path:
    """Pivot today's CSV log and write an Excel workbook with a line chart.

    Args:
        log_dir: directory that contains temps_YYYYMMDD.csv files
        threshold: alert threshold temperature (written as chart title annotation)

    Returns:
        Path to the generated .xlsx file

    Raises:
        FileNotFoundError: if today's log CSV does not exist
    """
    today = date.today()
    csv_path = log_dir / f"temps_{today:%Y%m%d}.csv"
    if not csv_path.exists():
        raise FileNotFoundError(csv_path)

    # ── Parse + pivot ──────────────────────────────────────────────────────
    # {timestamp_str: {projector_label: max_temp}}
    pivoted: dict[str, dict[str, float]] = defaultdict(dict)
    proj_order: list[str] = []  # preserve first-seen order

    with csv_path.open(newline="") as f:
        for row in csv.DictReader(f):
            if not row.get("temp_celsius"):
                continue
            ts = row["timestamp"]
            label = row["projector_label"]
            temp = float(row["temp_celsius"])
            if label not in proj_order:
                proj_order.append(label)
            existing = pivoted[ts].get(label)
            if existing is None or temp > existing:
                pivoted[ts][label] = temp

    if not pivoted:
        raise ValueError("No temperature readings found in today's log.")

    timestamps = sorted(pivoted.keys())

    # ── Build workbook ─────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Temps"

    # Header row
    hdr_fill = PatternFill("solid", fgColor="1A5C0F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Courier New")
    headers = ["Timestamp"] + proj_order
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    data_font = Font(name="Courier New", size=10)
    for row_idx, ts in enumerate(timestamps, start=2):
        # Format as HH:MM:SS
        try:
            ts_display = datetime.fromisoformat(ts).strftime("%H:%M:%S")
        except ValueError:
            ts_display = ts
        ws.cell(row=row_idx, column=1, value=ts_display).font = data_font
        for col_idx, label in enumerate(proj_order, start=2):
            temp = pivoted[ts].get(label)
            if temp is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=temp)
                cell.font = data_font
                cell.number_format = "0.0"

    # Column widths
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(proj_order) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    n_rows = len(timestamps)
    n_proj = len(proj_order)

    # ── Line chart ─────────────────────────────────────────────────────────
    chart = LineChart()
    chart.title = f"Projector Temperatures — {today:%Y-%m-%d}  (threshold {threshold:.0f}°C)"
    chart.style = 10
    chart.y_axis.title = "°C"
    chart.x_axis.title = "Time"
    chart.width = 30
    chart.height = 16

    # Data series — one per projector
    data_ref = Reference(ws, min_col=2, max_col=1 + n_proj,
                         min_row=1, max_row=1 + n_rows)
    chart.add_data(data_ref, titles_from_data=True)

    # Categories (timestamp labels)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_rows)
    chart.set_categories(cats)

    # Style each series
    for i, series in enumerate(chart.series):
        series.smooth = True
        color = _SERIES_COLORS[i % len(_SERIES_COLORS)]
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 20000  # 2 pt in EMUs

    # Place chart below the data table (leave a 2-row gap)
    chart_anchor_row = n_rows + 3
    ws.add_chart(chart, f"A{chart_anchor_row}")

    # ── Save ───────────────────────────────────────────────────────────────
    out_path = log_dir / f"report_{today:%Y%m%d}.xlsx"
    wb.save(out_path)
    return out_path
